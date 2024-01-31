"""
Creado por : Sacha Cruz
Fecha : 11/08/2023

Sistema de gestion industrial en Python
Programa realizado para gestiuonar en base a varios archivos Excel, una nueva tabla resumiendo sus valores.

---------------------------------------------------
Lista de las secciones de este codigo :
- Inicio
- Interfaz de Usuario
- Recuperar Excel
- Diccionario
- Añadir y calcular datos
- Bordes
- Estética
- Texto
- Colores
- Proteccion y Guardar el archivo
--------------------------------------------------
"""

from tkinter import Tk, Frame, Label, Button, LEFT, RIGHT, filedialog
from pandas import ExcelFile, read_excel
from os import path, system
from math import isnan, floor
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Side, Border, Font, Protection

# ___________________________________________________
# ---------------------------------------------------
# Interfaz de usuario
# ---------------------------------------------------

global excel_path

def open_file():
    global excel_path
    excel_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx")])
    if excel_path:
        root.destroy()

def open_excel_file():  
    try:
        system(f'start excel "{new_excel_path}"')
    except Exception as e:
        # print("Error:", e)
        pass
    root.destroy()

def close_window():
    root.destroy()
    
def tk_space():
    label = Label(root, text="")
    label.pack()
    
def tk_print(texto):
    label = Label(root, text = texto)
    Label()
    label.pack()

root = Tk()
root.title("Generación de un Resumen Mensual de Ventas")
root.attributes("-topmost", True)

screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
window_width = 450
window_height = 180
window_x_position = floor((screen_width - window_width) / 2)
window_y_position = floor((screen_height - window_height) / 2)
root.geometry(f"{window_width}x{window_height}+{window_x_position}+{window_y_position}")
tk_space()
tk_print("Seleccione el archivo Excel:")
tk_space()
open_button = Button(root, text="Abrir", command=open_file)
open_button.pack()
tk_space()
label_text_1 = "Si ya se ejecutó la aplicacion con su archivo, cierre la pagina Excel"
label_text_2 = "llamada 'RESUMEN MENSUAL + nombre' para que se actualice correctamente"
label = Label(root, text=label_text_1, font=("Helvetica", 8, "italic"))
label.pack()
label = Label(root, text=label_text_2, font=("Helvetica", 8, "italic"))
label.pack()
root.mainloop()

# ___________________________________________________
# ---------------------------------------------------
# Funciones
# ---------------------------------------------------

def round2(f):
    factor = 100 
    rounded_number = int(f * factor + 0.5) / factor
    return rounded_number

def showMatrix(m):
    """
    shows a
    m : matrix 
    """
    for row in m:
        for item in row:
            print(item, end="\t")
        print()  

def cleanMatrix(m):
    """
    cleans the matrix removing NaN values and useless lines
    m : matrix
    """
    total_matrix = []
    
    global ind_ref_tot 
    global ref
    ind_ref_tot = -1 # indice enlazado con la columna de valores totales
    ref = -1
    
    for line in m:
        if (type(line[0]) == str) :
            if ("total" in (line[0]).lower())  : # Solo guarda las lineas que tienen Total en su nombre
                ref +=1
                total_matrix.append(line)
                if (ind_ref_tot < 0) and ("dia" in (line[0]).lower()):
                    ind_ref_tot = ref
    for line in m:
        if (type(line[0]) == str) :
            if ("kilometraje" in (line[0]).lower())  : # Solo guarda la linea que tiene Kilometraje en su nombre
                total_matrix.append(line)
                
    matrix_h = len(total_matrix)
    matrix_w = len(total_matrix[0])
    total_matrix_cleaned = [ [] for k in range(matrix_h)]
    
    for i in range(matrix_h):
        (total_matrix_cleaned[i]).append(total_matrix[i][0])
        for j in range(1,matrix_w):
            a = total_matrix[i][j]
            if ((type(a) == int) or (type(a) == float)) and not(isnan(a)):
                a = round2(a)
                (total_matrix_cleaned[i]).append(a)
    return(total_matrix_cleaned)

def bsKilometro(m):
    """
    calcula la proporcion de Bs por kilometro por cada distribuidor
    retorna esa columna que luego hay que añadir a la matriz
    m : matrix
    """
    total_bs = []
    kilometraje = []
    bs_por_kilometro = ['BS. POR KILOMETRO'] + ["NaN" for k in range(6)]
    
    for line in m:
        if ("dia" in (line[0]).lower()) :
            total_bs = line
        if ("kilometraje" in (line[0]).lower()) :
            kilometraje = line
    for i in range(1, len(total_bs)):
        kilo = kilometraje[i]
        if kilo == 0 :
            bs_por_kilometro[i] = "---"
        else :
            bs_por_kilometro[i] = round2( (total_bs[i]) / (kilometraje[i]))
    return bs_por_kilometro

def colorFill(color, r1, r2, c1, c2):
    for row in ws.iter_rows(min_row = r1, max_row = r2, min_col = c1, max_col = c2):
        for cell in row:
            cell.fill = color

# ___________________________________________________
# ---------------------------------------------------
# Recuperacion del Excel
# ---------------------------------------------------

grupo_paginas_dias = []
grupo_paginas_limpias = []

try :
    with ExcelFile(excel_path) as xls:
        for sheet_name in xls.sheet_names:
            df = read_excel(xls, sheet_name)
            pagina_dia = df.values.tolist()[:100]
            grupo_paginas_dias.append(pagina_dia)
except FileNotFoundError :
    raise("Archivo no encontrado !")

# ___________________________________________________
# ---------------------------------------------------
# Limpiar los valores del excel
# ---------------------------------------------------

total_fecha = (grupo_paginas_dias[0])[1][8] 
total_year  = (grupo_paginas_dias[0])[1][5] 

# print(total_fecha)

for matrix in grupo_paginas_dias:
    matrix_cleaned = cleanMatrix(matrix)
    matrix_cleaned.insert(ind_ref_tot + 1, bsKilometro(matrix_cleaned))
    matrix_cleaned.insert(ind_ref_tot + 1, matrix_cleaned[-1])
    matrix_cleaned.pop(-1)
    grupo_paginas_limpias.append(matrix_cleaned)
    
# ___________________________________________________
# ---------------------------------------------------
# Inicio del diccionario
# ---------------------------------------------------

sample_matrix = grupo_paginas_limpias[0]

tabla_d = len(grupo_paginas_dias)    # Numero de dias
tabla_c = len(sample_matrix)         # Numero de categorias (Totales)
tabla_w = tabla_c*6 + 1              # LARGO de la tabla
tabla_h = tabla_d + 5                # ANCHO de la tabla
tabla_n = 6                          # Numero de distribuidores

linea_00 = ['REPORTE MENSUAL DE VENTAS - PROVEL LTDA.']
linea_01 = ['DIA']
linea_02 = ['DIA']


for k in range(len(sample_matrix)) :
    line = sample_matrix[k]
    if k == ind_ref_tot+1:
        title = line[0]
    elif k == ind_ref_tot+2:
        title = line[0] + " (Bs./Km)"
    else : 
        title = line[0] + " (Bs.)"
    linea_00 = linea_00 + ["" for k in range(tabla_n)]
    linea_01 = linea_01 + [title for k in range(tabla_n)]
    linea_02.append("D1")
    linea_02.append("D2")
    linea_02.append("D3")
    linea_02.append("D4")
    linea_02.append("D5")
    linea_02.append("D6")
    
linea_00[1:15] = ["FECHA : " + total_fecha + " " + str(total_year) for k in range(14)]


data = {'Titulo' : linea_00,
        'DIA': linea_01, 
        'Descripciones': linea_02}

# ___________________________________________________
# ---------------------------------------------------
# Añadir los datos
# ---------------------------------------------------

count = 0
for mat in grupo_paginas_limpias:
    count += 1
    linea_nueva = []
    for line in mat :
        linea_nueva = linea_nueva + line[1:]
    data[count] = [count] + linea_nueva

# ---------------------------------------------------
# Calculamos la suma de cada columna

total_suma = [0 for _ in range(tabla_w - 1)]
suma_parcial    = ["" for k in range(2)]
suma_porcentaje = ["" for k in range(2)]
texto_1      = ["" for k in range(2)]
texto_2      = ["" for k in range(2)]
texto_3      = ["" for k in range(2)]

for k in range(3, tabla_h-2):
    linea = data[list(data.keys())[k]][1:]
    for n in range(tabla_w-1):
        val = linea[n]
        if (type(val) == float) or (type(val) == int):
            total_suma[n] = total_suma[n] + linea[n]

# ---------------------------------------------------
# Calculamos las sumas parciales

for k in range(tabla_c):
    s = 0
    for val in total_suma[k*6:(k+1)*6] :
        s += val
    suma_parcial = suma_parcial + [s] + ["" for k in range(tabla_n-1)]
    if k < ind_ref_tot:
        texto_1 = texto_1 + ["Suma parcial :"] + ["" for k in range(tabla_n - 1)] 
        texto_2 = texto_2 + ["Porcentaje"] + ["" for k in range(tabla_n - 1)] 
        texto_3 = texto_3 + ["" for k in range(tabla_n)] 
    elif (k == ind_ref_tot+1):
        texto_1 = texto_1 + ["Total :"] + ["" for k in range(tabla_n - 1)] 
        texto_3 = texto_3 + ["" for k in range(6)] + ["Total (Valores > 0)"] + ["" for k in range(tabla_n - 1)] 
    elif (k == ind_ref_tot +2):
        texto_1 = texto_1 + ["Total :"] + ["" for k in range(tabla_n - 1)] 
        texto_3 = texto_3 + ["Promedio (Valores reales)"] + ["" for k in range(tabla_n - 1)] 
    else :
        texto_1 = texto_1 + ["Total :"] + ["" for k in range(tabla_n - 1)] 

# ---------------------------------------------------
# Calculamos los porcentajes

total_val = suma_parcial[ind_ref_tot*6 +2]

for k in range(tabla_c):
    s = 0
    for i in range(6):
        val = total_suma[k*6 + i]
        s += 100*(val/total_val)   
    suma_porcentaje = suma_porcentaje + [str(round(s, 2)) + " %"] + ["" for k in range(tabla_n - 1)] 
suma_porcentaje[ind_ref_tot*5+6 : ] = ["" for k in range(8)]

# ---------------------------------------------------
# Añadimos todo         

data["Total"]      = ["Suma Tot."] + total_suma
data["Space0"]     = ["" for k in range(tabla_w)]
data["Space1"]     = ["" for k in range(tabla_w)]
data["Space2"]     = [""] + texto_1
data["Parcial"]    = [""] + suma_parcial
data["Space3"]     = ["" for k in range(tabla_w)]
data["Space4"]     = [""] + texto_3
data["Porcentaje"] = [""] + suma_porcentaje
data["Space5"]     = ["" for k in range(tabla_w)]
data["Space6"]     = [""] + texto_2

# ---------------------------------------------------
# Calculamos promedios         

total_km     = data[list(data.keys())[tabla_h-2]][ind_ref_tot*7:ind_ref_tot*8-1]
total_bs_km  = data[list(data.keys())[tabla_h-2]][ind_ref_tot*8-1:ind_ref_tot*9-2]
prom_km      = []
prom_bs_km   = []
count = [0,0,0,0,0,0]

for k in range(3, tabla_h-2):
    linea_bs_km = data[list(data.keys())[k]][ind_ref_tot*8-1:ind_ref_tot*9-2]
    for j in range(tabla_n):
        val = linea_bs_km[j]
        t = type(val)
        if t ==int or t == float:
            count[j] += 1
s1 = 0
s2 = 0
cant = 0
for j in range(tabla_n):
    c = count[j]
    t1 = total_km[j]
    t2 = total_bs_km[j]
    if c == 0:
        prom_km.append( "---" )
        prom_bs_km.append( "---" )
    else:
        val_km = round(t1/c)
        val_bs_km = round(t2/c, 2)
        s1 += val_km
        s2 += val_bs_km
        cant += 1
        prom_km.append(val_km)
        prom_bs_km.append(val_bs_km)
promedio_km = round(s1/cant)
promedio_bs_km = round(s2/cant, 2)
        
(data["Total" ])[ind_ref_tot*7:ind_ref_tot*8-1] = prom_km
(data["Space1"])[ind_ref_tot*7:ind_ref_tot*8-1] = total_km
(data["Total" ])[ind_ref_tot*8-1 : ind_ref_tot*8+5] = prom_bs_km
(data["Space1"])[ind_ref_tot*8-1 : ind_ref_tot*8+5] = total_bs_km
(data["Space2"])[ind_ref_tot*7 : ind_ref_tot*8+5] = ["" for k in range(12)]
(data["Parcial"])[ind_ref_tot*8-1 : ind_ref_tot*8+5] = ["",""] + [promedio_bs_km] + ["", "", ""]
                                                                         
# ___________________________________________________
# ---------------------------------------------------
# Poner todo en una nueva hoja excel
# ---------------------------------------------------

wb = Workbook()
ws = wb.active

for row_data in data.values():
    ws.append(row_data)

# ___________________________________________________
# ---------------------------------------------------
# Bordes
# ---------------------------------------------------

border_side_1 = Side(style="thin", color="000000")  
border_side_2 = Side(style="thin", color="000000")  
border_side_3 = Side(border_style="thick", color="000000")
border_side_4 = Side(border_style="double", color="000000")
border_side_5 = Side(border_style="dotted", color="000000")

# bordes verticales
for row in ws.iter_rows(min_row=1, max_row=tabla_h, min_col=1, max_col=tabla_w):
    for cell in row:
        cell.border = Border(left=border_side_1, right=border_side_1)
for row in ws.iter_rows(min_row=2, max_row=tabla_h-1, min_col=1, max_col=1):
    for cell in row:
        cell.border = Border(right = border_side_2)
# bordes verticales por cada categoria
for k in range(tabla_c+1):
    for row in ws.iter_rows(min_row=1, max_row=tabla_h, min_col=1+6*k, max_col=1+6*k):
        for cell in row:
            cell.border = Border(right=border_side_4)
# bordes alrededor de totales y kilometraje
for row in ws.iter_rows(min_row=2, max_row=tabla_h-1, min_col= ind_ref_tot*8 + 5,  max_col= ind_ref_tot*8 + 5):
    for cell in row:
        cell.border = Border(right = border_side_3)
for row in ws.iter_rows(min_row=2, max_row=tabla_h-1, min_col= ind_ref_tot*6 +1,  max_col= ind_ref_tot*6 +1):
    for cell in row:
        cell.border = Border(right = border_side_3)
# bordes horizontales arriba y abajo
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=tabla_w):
    for cell in row:
        cell.border = Border(bottom = border_side_2)
for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=tabla_w):
    for cell in row:
        cell.border = Border(bottom = border_side_2)
for row in ws.iter_rows(min_row=tabla_h-1, max_row=tabla_h-1, min_col=1, max_col=tabla_w):
    for cell in row:
        cell.border = Border(bottom = border_side_2, top = border_side_2)
# bordes de las casillas inferiores
for k in range(tabla_c):
    if k == ind_ref_tot:
        for row in ws.iter_rows(min_row=tabla_h+3, max_row=tabla_h+4, min_col= (4 + 6*k), max_col = (5 + 6*k)) :
            for cell in row:
                cell.border = Border(left=border_side_3, right=border_side_3, top =border_side_3, bottom = border_side_3)
    else:
        for row in ws.iter_rows(min_row=tabla_h+3, max_row=tabla_h+4, min_col= (4 + 6*k), max_col = (5 + 6*k)) :
            for cell in row:
                cell.border = Border(left=border_side_2, right=border_side_2, top =border_side_2, bottom = border_side_2)
        if k <= ind_ref_tot-1:
            for row in ws.iter_rows(min_row=tabla_h+6, max_row=tabla_h+7, min_col= (4 + 6*k), max_col = (5 + 6*k)) :
                for cell in row:
                    cell.border = Border(left=border_side_2, right=border_side_2, top =border_side_2, bottom = border_side_2)
# linea horizontal de delimitacion :
for row in ws.iter_rows(min_row=tabla_h+9, max_row=tabla_h+9, min_col=1, max_col =tabla_w) :
    for cell in row:
        cell.border = Border(top = border_side_5)
# linea vertical de delimitacion :
for row in ws.iter_rows(min_row=tabla_h+1, max_row=tabla_h+8, min_col = tabla_w+1, max_col = tabla_w+1) :
    for cell in row:
        cell.border = Border(left = border_side_5)
# Corregir esquinas abajo
for k in range(tabla_c+1):
    for row in ws.iter_rows(min_row=tabla_h, max_row=tabla_h, min_col= (1 + 6*k), max_col = (1 + 6*k)) :
        for cell in row:
            cell.border = Border(right=border_side_4, top =border_side_2, bottom = border_side_2)
# Corregir esquinas arriba
for k in range(tabla_c+1):
    for row in ws.iter_rows(min_row=2, max_row=2, min_col= (1 + 6*k), max_col = (1 + 6*k)) :
        for cell in row:
            cell.border = Border(right=border_side_4, top =border_side_2, bottom = border_side_2)
# bordes totales debajo de promedio
for row in ws.iter_rows(min_row=tabla_h+1, max_row=tabla_h+1, min_col = ind_ref_tot*7+1, max_col = ind_ref_tot*8+5) :
    for cell in row:
        cell.border = Border(border_side_2, border_side_2, border_side_2, border_side_2)

# ___________________________________________________
# ---------------------------------------------------
# Estética
# ---------------------------------------------------

# ---------------------------------------------------
# AGRANDAR CELDAS

# Alargar todas las celdas
for column in ws.columns:
    ws.column_dimensions[column[0].column_letter].width = 10

# Hacer mas alta la primera linea
ws.row_dimensions[1].height = 35

# ---------------------------------------------------
# UNIR CELDAS

# Titulo
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
ws.merge_cells(start_row=1, start_column=11, end_row=1, end_column=15)
ws.merge_cells(start_row=1, start_column=16, end_row=1, end_column=tabla_w)
# Dia
ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
# Promedio texto
ws.merge_cells(start_row=tabla_h+5, end_row=tabla_h+5, start_column=ind_ref_tot*7+3, end_column=ind_ref_tot*7+4)
ws.merge_cells(start_row=tabla_h+5, end_row=tabla_h+5, start_column=ind_ref_tot*7+9, end_column=ind_ref_tot*7+10)
# Suma Tot.
for i in range(1, tabla_w+1):
    ws.merge_cells(start_row=tabla_h-1, start_column=i, end_row=tabla_h, end_column=i)
# Unir las celdas periodicas
for k in range(tabla_c):
    # Titulos
    ws.merge_cells(start_row=2, start_column= (2 + 6*k), end_row=2, end_column= (7 + 6*k))
    # Nombre de los totales abajo
    ws.merge_cells(start_row=tabla_h+3, start_column= (4 + 6*k), end_row=tabla_h+4, end_column= (5 + 6*k))
    if k < ind_ref_tot:
        ws.merge_cells(start_row=tabla_h+8, start_column= (4 + 6*k), end_row=tabla_h+8, end_column= (5 + 6*k))
        # Porcentajes abajo
        ws.merge_cells(start_row=tabla_h+6, start_column= (4 + 6*k), end_row=tabla_h+7, end_column= (5 + 6*k))
    # Totales abajo
    ws.merge_cells(start_row=tabla_h+2, start_column= (4 + 6*k), end_row=tabla_h+2, end_column= (5 + 6*k))
        
# ___________________________________________________
# ---------------------------------------------------
# TEXTO
# ---------------------------------------------------

small_font = Font(size=9)
mid_font = Font(size=10.5)
bold_font = Font(bold=True, italic=False, underline='none', color="000000")
mid_bold_font = Font(size=12, bold=True, italic=False, underline='none', color="000000")
big_bold_font = Font(bold=True, size=16)

# Achicar los textos de abajo
for row in ws.iter_rows(min_row=tabla_h+1, max_row=tabla_h+8, min_col=1, max_col=tabla_w):
    for cell in row:
        cell.font = small_font
# Agrandar los titulos
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=tabla_w):
    for cell in row:
        cell.font = mid_bold_font
# Poner en negrita los titulos
for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=tabla_w):
    for cell in row:
        cell.font = bold_font
# Poner en negrita los totales
for row in ws.iter_rows(min_row=tabla_h-1, max_row=tabla_h+1, min_col=2, max_col=tabla_w):
    for cell in row:
        cell.font = mid_bold_font
for row in ws.iter_rows(min_row=tabla_h+3, max_row=tabla_h+6, min_col=2, max_col=tabla_w):
    for cell in row:
        cell.font = big_bold_font
# Añadir 2 decimales a cada valor en Bs.
for row in ws.iter_rows(min_row=3, max_row=tabla_h+10, min_col=2, max_col=tabla_w):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0.00'
# Excepto por los kilometrajes
for row in ws.iter_rows(min_row=3, max_row=tabla_h+5, min_col=ind_ref_tot*7 +1, max_col=ind_ref_tot*7+6):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = '0'
# Centrar todo el texto
for row in ws.iter_rows(min_row=1, max_row=tabla_h+10, min_col=1, max_col=100):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
# Pero las cifras a la derecha
for row in ws.iter_rows(min_row=4, max_row=tabla_h+1, min_col=2, max_col=tabla_w):
    for cell in row:
        cell.alignment = Alignment(horizontal='right', vertical='center')
# Texto de promedios
for row in ws.iter_rows(min_row=tabla_h+5, max_row=tabla_h+5, min_col=2, max_col=tabla_w):
    for cell in row:
        cell.font = small_font
# ___________________________________________________
# ---------------------------------------------------
# Colores
# ---------------------------------------------------

yellow_fill         = PatternFill(start_color="fffcc7", end_color="fffcc7", fill_type="solid")
turquoise_fill      = PatternFill(start_color="c5fcd6", end_color="c5fcd6", fill_type="solid")
white_fill          = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
blue_fill           = PatternFill(start_color="a5b5fa", end_color="a5b5fa", fill_type="solid")
light_blue_fill     = PatternFill(start_color="bfc9f2", end_color="bfc9f2", fill_type="solid")
light_gray_fill     = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
lighter_gray_fill   = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
pale_yellow_fill    = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

#detalles
colorFill(light_blue_fill, 1, 1, 1, 20)
colorFill(blue_fill, 1, 1, 11, 15)
colorFill(light_blue_fill, 1, 1, 16, tabla_w)
colorFill(pale_yellow_fill, 2, 2, 2, tabla_w)
colorFill(blue_fill, 2, tabla_h, 1, 1)

# Pintar zebrado vertical
kmax = int(tabla_d/2+1)
for k in range(kmax):
    colorFill(lighter_gray_fill, 2*k+4, 2*k+4, 2, tabla_w)
    colorFill(light_gray_fill, 2*k+4, 2*k+4, 1, 1)
    if k<kmax-1:
        colorFill(white_fill, 2*k+5, 2*k+5, 2, tabla_w)
        colorFill(lighter_gray_fill, 2*k+5, 2*k+5, 1, 1)
# Pintar distribuidores y total abajo
for k in range(int(tabla_w/2)):
    colorFill(yellow_fill, 3, 3, 2+2*k, 2+2*k)
    colorFill(turquoise_fill, 3, 3, 3+2*k, 3+2*k)
    if (k < ind_ref_tot*7/2-1) or (k > (ind_ref_tot*8+4)/2-1):
        colorFill(turquoise_fill, tabla_h-1, tabla_h, 3+2*k, 3+2*k)
        colorFill(yellow_fill, tabla_h-1, tabla_h, 2+2*k, 2+2*k)
    else :
        colorFill(blue_fill , tabla_h-1, tabla_h, 3+2*k, 3+2*k)
        colorFill(light_blue_fill , tabla_h-1, tabla_h, 2+2*k, 2+2*k)
        colorFill(turquoise_fill, tabla_h, tabla_h+1, 3+2*k, 3+2*k)
        colorFill(yellow_fill, tabla_h, tabla_h+1, 2+2*k, 2+2*k)

# ___________________________________________________
# ---------------------------------------------------
# Proteccion y guardar el archivo
# ---------------------------------------------------

ws.protection.sheet = True

for row in ws.iter_rows(min_row=tabla_h+9, max_row=tabla_h+40, min_col=1, max_col=tabla_w+40    ):
    for cell in row:
        cell.protection = Protection(locked=False)

# Get the name and extension of the original Excel file
excel_name, _ = path.splitext(path.basename(excel_path))
new_excel_name = f"RESUMEN MENSUAL {excel_name}.xlsx"
new_excel_path = path.join(path.dirname(excel_path), new_excel_name)

excel_exists = path.exists(new_excel_path)

# ___________________________________________________
# ---------------------------------------------------
# Interfaz
# ---------------------------------------------------

root = Tk()
root.attributes("-topmost", True)
root.geometry(f"{550}x{100}+{floor((root.winfo_screenwidth() - 550) / 2)}+{floor((root.winfo_screenheight() - 100) / 2)}")

try:
    wb.save(new_excel_path)
    root.title("Archivo actualizado !" if excel_exists else "Nuevo archivo creado !")
    tk_space()
    tk_print(f"El archivo Excel {new_excel_name} {'ya existente fue actualizado !' if excel_exists else 'fue creado exitosamente !'}")    
    tk_space()
    button_frame = Frame(root)
    button_frame.pack()
    Button(button_frame, text="Abrir Excel", command=open_excel_file).pack(side=LEFT, padx=10)
    Button(button_frame, text="Ok", command=close_window).pack(side=RIGHT, padx=10)
    root.mainloop()
    
except PermissionError:
    root.title("Por favor cierre el archivo Excel antes de ejecutar")
    tk_space()
    tk_print(f"Por favor cierre el archivo {new_excel_name} antes de ejecutar !")
    tk_space()
    tk_print("No se olvide de guardar los cambios que hizo a la hoja bajo otro nombre")
    tk_space()
    Button(root, text="Ok", command=close_window).pack()
    root.mainloop()